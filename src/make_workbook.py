"""Build a 12-tab monthly bill workbook from the repository's own calendar.

One tab per month, chronological, covering the next full 12 months. Every date
and every amount comes from `bills.json` / `income.json` through `fincal`, which
is the same authority the forecast and the dashboard read — so the workbook
cannot drift from the engine.

Two amount columns, deliberately. "Budget" is what the engine forecasts for that
occurrence; "Stored" is the figure recorded in bills.json. They differ in three
documented cases, and a workbook showing only one of them would hide the reason:

  * a seasonal bill carries a month-of-year profile, so its budget figure moves
    across the year and neither number is wrong;
  * a variable bill landing monthly or less often is forecast at the top of its
    observed range, which is the deliberate forecast-at-worst rule;
  * four cards and the water bill are forecast ABOVE an amount the household
    renegotiated, because `observed_max` predates the correction. That one is a
    defect, flagged SUPERSEDED in the Basis column.

    python src/make_workbook.py --out monthly-bills.xlsx
"""

from __future__ import annotations

import argparse
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import fincal
from bills import load_items, money

ROOT = Path(__file__).resolve().parent.parent

FONT = "Arial"
MONEY_FMT = '$#,##0.00;($#,##0.00);-'

NAVY = "17365D"
HEAD_FILL = PatternFill("solid", fgColor="DCE3EA")
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
TOTAL_FILL = PatternFill("solid", fgColor="EDF1F5")
WARN_FILL = PatternFill("solid", fgColor="FFF3CD")
BAD_FILL = PatternFill("solid", fgColor="FBE4E4")
GOOD_FILL = PatternFill("solid", fgColor="E6F2EA")
INPUT_BLUE = Font(name=FONT, size=10, color="0000FF")
THIN = Side(style="thin", color="BFC7CF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TIER_LABEL = {1: "1 Secured/critical", 2: "2 Essential", 3: "3 Required debt",
              4: "4 Scheduled", 5: "5 Discretionary"}

BILL_COLS = [
    ("Due date", 12), ("Day", 6), ("Bill / payee", 26), ("Category", 18),
    ("Frequency", 11), ("Budget", 12), ("Stored", 12), ("Basis", 22),
    ("Movability", 13), ("Autopay", 9), ("Historical range seen", 22),
]
INC_COLS = [
    ("Date", 12), ("Day", 6), ("Source", 26), ("Frequency", 11),
    ("Amount", 12), ("Status", 30),
]


def month_span(first: date, n: int) -> list[tuple[int, int]]:
    out, y, m = [], first.year, first.month
    for _ in range(n):
        out.append((y, m))
        m += 1
        if m == 13:
            y, m = y + 1, 1
    return out


def month_bounds(y: int, m: int) -> tuple[date, date]:
    start = date(y, m, 1)
    end = (date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)) - timedelta(days=1)
    return start, end


def basis_of(b: dict, budget, stored) -> tuple[str, str]:
    """Why the budget figure is what it is, and how much to shout about it."""
    if b.get("monthly_expected"):
        return "Seasonal profile", "info"
    if budget == stored:
        return "Stated amount", "ok"
    if b.get("source") == "household":
        return "SUPERSEDED - see notes", "bad"
    return "Forecast at worst seen", "warn"


def build(out: Path, start: date, months: int = 12) -> tuple[int, dict]:
    bills = load_items(ROOT / "bills.json", "bills")
    income = load_items(ROOT / "income.json", "income")
    active = [b for b in bills if b.get("active", True)]
    inactive = [b for b in bills if not b.get("active", True)]
    off_income = [i for i in income if not i.get("active", True)]

    wb = Workbook()
    wb.remove(wb.active)
    totals: dict[str, tuple[float, float]] = {}

    for y, m in month_span(start, months):
        lo, hi = month_bounds(y, m)
        label = lo.strftime("%b %Y")
        ws = wb.create_sheet(label)

        # ---- title -------------------------------------------------------
        ws.merge_cells("A1:K1")
        c = ws["A1"]
        c.value = f"OpenFIN  -  Monthly bill plan  -  {lo.strftime('%B %Y')}"
        c.font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
        c.fill = TITLE_FILL
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 26

        ws["A2"] = ("Dates and amounts generated from bills.json and income.json. "
                    "Seasonal bills use their month-of-year profile built from "
                    "recent years of statements.")
        ws["A2"].font = Font(name=FONT, size=9, italic=True, color="5A6570")
        ws.merge_cells("A2:K2")

        # ---- income ------------------------------------------------------
        r = 4
        ws.cell(r, 1, "MONEY IN").font = Font(name=FONT, size=11, bold=True,
                                              color=NAVY)
        r += 1
        for j, (t, w) in enumerate(INC_COLS, start=1):
            cell = ws.cell(r, j, t)
            cell.font = Font(name=FONT, size=10, bold=True)
            cell.fill = HEAD_FILL
            cell.border = BOX
            ws.column_dimensions[get_column_letter(j)].width = max(
                w, ws.column_dimensions[get_column_letter(j)].width or 0)
        inc_head = r
        r += 1
        inc_first = r
        rows = []
        for i in income:
            if not i.get("active", True):
                continue
            for d in fincal.occurrences(i, lo, hi):
                rows.append((d, i))
        for d, i in sorted(rows, key=lambda x: x[0]):
            ws.cell(r, 1, d).number_format = "dd mmm yyyy"
            ws.cell(r, 2, d.strftime("%a"))
            ws.cell(r, 3, i["name"])
            ws.cell(r, 4, i.get("frequency", ""))
            amt = ws.cell(r, 5, float(money(i["amount"])))
            amt.number_format = MONEY_FMT
            amt.font = INPUT_BLUE
            ws.cell(r, 6, "UNVERIFIED - flagged needs_review"
                    if i.get("needs_review") else "Confirmed from statements")
            for j in range(1, 7):
                ws.cell(r, j).border = BOX
                if ws.cell(r, j).font.color is None or j != 5:
                    ws.cell(r, j).font = Font(name=FONT, size=10)
            r += 1
        inc_last = r - 1
        if inc_last < inc_first:                     # no income at all
            ws.cell(r, 3, "None modelled this month")
            ws.cell(r, 3).font = Font(name=FONT, size=10, italic=True)
            inc_first, inc_last, r = r, r, r + 1
        inc_total_row = r
        ws.cell(r, 3, "Total money in").font = Font(name=FONT, size=10, bold=True)
        tc = ws.cell(r, 5, f"=SUM(E{inc_first}:E{inc_last})")
        tc.number_format = MONEY_FMT
        tc.font = Font(name=FONT, size=10, bold=True)
        for j in range(1, 7):
            ws.cell(r, j).fill = TOTAL_FILL
            ws.cell(r, j).border = BOX

        # ---- bills -------------------------------------------------------
        r += 2
        ws.cell(r, 1, "MONEY OUT").font = Font(name=FONT, size=11, bold=True,
                                               color=NAVY)
        r += 1
        for j, (t, w) in enumerate(BILL_COLS, start=1):
            cell = ws.cell(r, j, t)
            cell.font = Font(name=FONT, size=10, bold=True)
            cell.fill = HEAD_FILL
            cell.border = BOX
            ws.column_dimensions[get_column_letter(j)].width = w
        head_row = r
        r += 1
        first = r

        occ = []
        for b in active:
            for d in fincal.occurrences(b, lo, hi):
                occ.append((d, b))
        for d, b in sorted(occ, key=lambda x: (x[0], -float(
                fincal.expected_amount(x[1], x[0])))):
            budget = fincal.expected_amount(b, d)
            stored = money(b["amount"])
            basis, tone = basis_of(b, budget, stored)
            secured = b.get("secured")
            movable = ("FIXED" if secured or not b.get("deferrable", False)
                       else "UNKNOWN")
            rng = ""
            if b.get("observed_min") is not None or b.get("observed_max") is not None:
                rng = (f"{money(b.get('observed_min', 0)):,.2f}"
                       f" - {money(b.get('observed_max', 0)):,.2f}")
            elif b.get("monthly_expected"):
                vals = [money(v) for v in b["monthly_expected"].values()]
                rng = f"{min(vals):,.2f} - {max(vals):,.2f} across the year"

            ws.cell(r, 1, d).number_format = "dd mmm yyyy"
            ws.cell(r, 2, d.strftime("%a"))
            ws.cell(r, 3, b["name"])
            ws.cell(r, 4, TIER_LABEL.get(b.get("priority_tier", 5), "5"))
            ws.cell(r, 5, b.get("frequency", ""))
            bc = ws.cell(r, 6, float(budget))
            bc.number_format = MONEY_FMT
            bc.font = Font(name=FONT, size=10, color="0000FF", bold=True)
            sc = ws.cell(r, 7, float(stored))
            sc.number_format = MONEY_FMT
            sc.font = INPUT_BLUE
            ws.cell(r, 8, basis)
            ws.cell(r, 9, movable)
            ws.cell(r, 10, "yes" if b.get("autopay") else "no")
            ws.cell(r, 11, rng)
            fill = {"bad": BAD_FILL, "warn": WARN_FILL}.get(tone)
            for j in range(1, 12):
                cell = ws.cell(r, j)
                cell.border = BOX
                if j not in (6, 7):
                    cell.font = Font(name=FONT, size=10)
                if fill and j == 8:
                    cell.fill = fill
                if j == 9 and movable == "FIXED":
                    cell.font = Font(name=FONT, size=10, bold=True)
            r += 1
        last = r - 1

        total_row = r
        ws.cell(r, 3, "Total money out").font = Font(name=FONT, size=10, bold=True)
        for col, letter in ((6, "F"), (7, "G")):
            t = ws.cell(r, col, f"=SUM({letter}{first}:{letter}{last})")
            t.number_format = MONEY_FMT
            t.font = Font(name=FONT, size=10, bold=True)
        for j in range(1, 12):
            ws.cell(r, j).fill = TOTAL_FILL
            ws.cell(r, j).border = BOX

        # ---- summary -----------------------------------------------------
        r += 2
        ws.cell(r, 1, "MONTH SUMMARY").font = Font(name=FONT, size=11, bold=True,
                                                   color=NAVY)
        r += 1
        summary = [
            ("Money in", f"=E{inc_total_row}"),
            ("Money out (budget)", f"=F{total_row}"),
            ("Net for the month", f"=E{inc_total_row}-F{total_row}"),
            ("Money out at stored amounts", f"=G{total_row}"),
            ("Over-forecast vs stored", f"=F{total_row}-G{total_row}"),
        ]
        for lbl, formula in summary:
            ws.cell(r, 3, lbl).font = Font(name=FONT, size=10,
                                           bold=lbl.startswith("Net"))
            v = ws.cell(r, 6, formula)
            v.number_format = MONEY_FMT
            v.font = Font(name=FONT, size=10, bold=lbl.startswith("Net"))
            for j in (3, 4, 5, 6):
                ws.cell(r, j).border = BOX
                if lbl.startswith("Net"):
                    ws.cell(r, j).fill = TOTAL_FILL
            r += 1

        # ---- notes -------------------------------------------------------
        r += 1
        ws.cell(r, 1, "NOTES AND ASSUMPTIONS").font = Font(
            name=FONT, size=11, bold=True, color=NAVY)
        r += 1
        notes = [
            "Budget vs Stored: 'Budget' is what the engine forecasts for that "
            "date; 'Stored' is the figure recorded in bills.json.",
            "Seasonal profile: gas and electric carry a month-of-year profile "
            "built from recent years of statements, so their budget figure "
            "changes every month. Neither figure is wrong.",
            "Forecast at worst seen: a variable bill landing monthly or less "
            "often is budgeted at the top of its observed range, deliberately, "
            "because under-forecasting an outflow is what causes an overdraft.",
            "SUPERSEDED: the engine is forecasting ABOVE an amount the "
            "household renegotiated, because the observed maximum predates the "
            "correction. Affects four cards and the water bill. This is a "
            "defect, not a rule - the Stored column is the agreed figure.",
            "Movability: FIXED means secured or recorded as non-deferrable. "
            "UNKNOWN means nothing records how late the payment may safely be, "
            "so no date change can be recommended without confirming it.",
            "Weekly items (groceries, fuel) appear once per occurrence, so a "
            "five-week month legitimately shows five rows.",
        ]
        if off_income:
            notes.insert(0, (
                "INCOME SWITCHED OFF: "
                + "; ".join(f"{i['name']} ({money(i['amount']):,.2f} "
                            f"{i.get('frequency')})" for i in off_income)
                + " is recorded in income.json but NOT counted anywhere in this "
                  "workbook. If it is actually being received, every monthly "
                  "net figure here is understated by that amount."))
        if inactive:
            notes.append(
                "Excluded as ended or paid off, kept for history: "
                + ", ".join(b["name"] for b in inactive) + ".")
        for n in notes:
            ws.cell(r, 1, "- " + n).font = Font(name=FONT, size=9,
                                                color="42474C")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
            ws.cell(r, 1).alignment = Alignment(wrap_text=False)
            r += 1

        ws.freeze_panes = ws.cell(head_row + 1, 1)
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = f"{head_row}:{head_row}"
        totals[label] = (inc_total_row, total_row)

    wb.save(out)
    return len(wb.sheetnames), totals


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="12-tab monthly bill workbook")
    ap.add_argument("--out", default="monthly-bills.xlsx")
    ap.add_argument("--start", help="first month, YYYY-MM (default: next month)")
    ap.add_argument("--months", type=int, default=12)
    a = ap.parse_args(argv)

    if a.start:
        y, m = (int(x) for x in a.start.split("-"))
        first = date(y, m, 1)
    else:
        t = date.today()
        first = date(t.year + (t.month == 12), 1 if t.month == 12 else t.month + 1, 1)

    n, _ = build(Path(a.out), first, a.months)
    print(f"{a.out} written - {n} tabs, {first:%b %Y} onward.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
